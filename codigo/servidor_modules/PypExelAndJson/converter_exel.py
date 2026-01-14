import json
import pandas as pd
import base64
import tempfile
import os
from typing import Dict, Any, List, Optional
from datetime import datetime
from openpyxl import Workbook
import traceback

class ExcelConverter:
    """Conversor completo para Excel ↔ JSON com suporte a TUBOS"""
    
    def excel_to_json(self, excel_bytes: bytes, filename: str) -> Dict[str, Any]:
        """
        Converte arquivo Excel para JSON estruturado do sistema
        """
        try:
            print(f"🔄 Iniciando conversão Excel para JSON: {filename}")
            
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
                f.write(excel_bytes)
                temp_path = f.name
                print(f"📄 Arquivo temporário criado: {temp_path}")
            
            try:
                # Ler arquivo Excel
                with pd.ExcelFile(temp_path) as xls:
                    result = {
                        "constants": {},
                        "machines": [],
                        "materials": {},
                        "empresas": [],
                        "banco_equipamentos": {},
                        "dutos": [],
                        "tubos": []
                    }
                    
                    print(f"📊 Sheets encontradas: {xls.sheet_names}")
                    
                    # Processar cada sheet
                    for sheet_name in xls.sheet_names:
                        try:
                            df = pd.read_excel(xls, sheet_name=sheet_name)
                            sheet_name_lower = sheet_name.lower()
                            
                            print(f"📋 Processando sheet: '{sheet_name}' ({len(df)} linhas)")
                            
                            # Identificar tipo de sheet
                            if 'constant' in sheet_name_lower or sheet_name_lower == 'constants':
                                self._process_constants_sheet(df, result)
                            elif 'machine' in sheet_name_lower or sheet_name_lower == 'machines':
                                self._process_machines_sheet(df, result)
                            elif 'material' in sheet_name_lower or sheet_name_lower == 'materials':
                                self._process_materials_sheet(df, result)
                            elif 'empresa' in sheet_name_lower or sheet_name_lower == 'empresas':
                                self._process_empresas_sheet(df, result)
                            elif 'equipamento' in sheet_name_lower or 'banco_equipamentos' in sheet_name_lower:
                                self._process_equipamentos_sheet(df, result)
                            elif 'duto' in sheet_name_lower or sheet_name_lower == 'dutos':
                                self._process_dutos_sheet(df, result)
                            elif 'tubo' in sheet_name_lower or sheet_name_lower == 'tubos':
                                self._process_tubos_sheet(df, result)
                            elif 'pipe' in sheet_name_lower or 'tubes' in sheet_name_lower:
                                self._process_tubos_sheet(df, result)
                            else:
                                print(f"⚠️  Sheet '{sheet_name}' não reconhecida, ignorando...")
                                
                        except Exception as sheet_error:
                            print(f"⚠️  Erro no sheet '{sheet_name}': {sheet_error}")
                            traceback.print_exc()
                            continue
                    
                    # Validação final
                    validation = self._validate_structure(result)
                    if not validation["valid"]:
                        error_msg = f"Estrutura inválida: {', '.join(validation['errors'])}"
                        print(f"❌ {error_msg}")
                        return {
                            "success": False,
                            "error": error_msg
                        }
                    
                    # Resumo da conversão
                    print("\n" + "="*50)
                    print("✅ CONVERSÃO EXCEL → JSON CONCLUÍDA")
                    print("="*50)
                    print(f"📊 Constants: {len(result['constants'])}")
                    print(f"⚙️  Machines: {len(result['machines'])}")
                    print(f"📦 Materials: {len(result['materials'])}")
                    print(f"🏢 Empresas: {len(result['empresas'])}")
                    print(f"🔧 Equipamentos: {len(result['banco_equipamentos'])} tipos")
                    print(f"🟦 Dutos: {len(result['dutos'])} tipos")
                    print(f"🔵 Tubos: {len(result['tubos'])} tipos")
                    print("="*50)
                    
                    return {
                        "success": True,
                        "data": result,
                        "message": f"Arquivo '{filename}' convertido com sucesso",
                        "metadata": {
                            "filename": filename,
                            "sheets_processed": xls.sheet_names,
                            "structure_valid": True
                        }
                    }
                    
            finally:
                # Limpar arquivo temporário
                if os.path.exists(temp_path):
                    try:
                        os.unlink(temp_path)
                        print(f"🧹 Arquivo temporário removido: {temp_path}")
                    except Exception as e:
                        print(f"⚠️  Não foi possível remover temp file: {e}")
                        
        except Exception as e:
            print(f"❌ Erro crítico ao converter Excel: {e}")
            traceback.print_exc()
            return {
                "success": False,
                "error": f"Erro na conversão: {str(e)}"
            }
    
    def _process_constants_sheet(self, df: pd.DataFrame, result: Dict[str, Any]) -> None:
        """Processa sheet de constantes"""
        try:
            print(f"🔢 Processando constants ({len(df)} linhas)")
            
            # Determinar se tem cabeçalho
            has_header = False
            if len(df) > 0:
                first_row = df.iloc[0]
                if any(isinstance(val, str) and val.lower() in ['key', 'chave', 'nome'] for val in first_row.values[:3]):
                    has_header = True
            
            start_row = 1 if has_header else 0
            
            for idx in range(start_row, len(df)):
                row = df.iloc[idx]
                
                # Pular linhas vazias
                if len(row) < 1 or pd.isna(row.iloc[0]):
                    continue
                
                key = str(row.iloc[0]).strip()
                if not key:
                    continue
                
                # Encontrar valor (coluna 2 ou primeira não vazia após key)
                value = None
                value_col = None
                
                for col_idx in range(1, min(5, len(row))):
                    if pd.notna(row.iloc[col_idx]) and str(row.iloc[col_idx]).strip() != '':
                        value = row.iloc[col_idx]
                        value_col = col_idx
                        break
                
                if value is None:
                    print(f"⚠️  Linha {idx}: Key '{key}' sem valor, pulando...")
                    continue
                
                # Descrição (coluna após valor)
                description = ""
                if value_col is not None and value_col + 1 < len(row) and pd.notna(row.iloc[value_col + 1]):
                    description = str(row.iloc[value_col + 1]).strip()
                
                # Converter valor
                try:
                    if isinstance(value, (int, float)):
                        num_value = float(value)
                    elif isinstance(value, str):
                        # Tentar converter string para número
                        try:
                            num_value = float(value.replace(',', '.'))
                        except:
                            num_value = value
                    else:
                        num_value = value
                except:
                    num_value = value
                
                # Adicionar à estrutura
                result["constants"][key] = {
                    "value": num_value,
                    "description": description,
                    "unit": "",
                    "category": ""
                }
                
                print(f"  ✅ {key}: {num_value} ({description[:30]})")
            
            print(f"✅ Constants processados: {len(result['constants'])}")
            
        except Exception as e:
            print(f"⚠️  Erro ao processar constants: {e}")
            traceback.print_exc()
    
    def _process_machines_sheet(self, df: pd.DataFrame, result: Dict[str, Any]) -> None:
        """Processa sheet de máquinas"""
        try:
            print(f"⚙️  Processando machines ({len(df)} linhas)")
            
            # Detectar formato
            has_detailed_format = False
            if len(df.columns) >= 5:
                # Verificar se tem colunas detalhadas
                col_names = [str(col).lower() for col in df.columns[:5]]
                if any(name in col_names for name in ['type', 'descrição', 'description', 'categoria', 'peso']):
                    has_detailed_format = True
            
            if has_detailed_format:
                # Formato detalhado (colunas nomeadas)
                print("📋 Formato detalhado detectado")
                for idx, row in df.iterrows():
                    if idx == 0 and any('type' in str(col).lower() for col in df.columns):
                        continue  # Pular cabeçalho
                    
                    machine_type = row.iloc[0] if len(row) > 0 else None
                    if pd.isna(machine_type) or str(machine_type).strip() == '':
                        continue
                    
                    machine = {
                        "type": str(machine_type).strip(),
                        "description": str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else "",
                        "impostos": {
                            "PIS_COFINS": "INCL",
                            "IPI": "ISENTO",
                            "ICMS": "12%",
                            "PRAZO": "45 a 60 dias",
                            "FRETE": "FOB/Cabreúva/SP"
                        },
                        "configuracoes_instalacao": [],
                        "baseValues": {},
                        "options": [],
                        "voltages": [],
                        "dimensions": {},
                        "peso": row.iloc[3] if len(row) > 3 and pd.notna(row.iloc[3]) else 0,
                        "categoria": str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else ""
                    }
                    
                    result["machines"].append(machine)
                    print(f"  ✅ Máquina: {machine['type']}")
            else:
                # Formato simples (tipo, capacidade, valor)
                print("📋 Formato simples detectado")
                current_machine = None
                
                for idx, row in df.iterrows():
                    if len(row) < 1 or pd.isna(row.iloc[0]):
                        continue
                    
                    first_cell = str(row.iloc[0]).strip()
                    
                    # Nova máquina
                    if 'type:' in first_cell.lower() or 'tipo:' in first_cell.lower() or current_machine is None:
                        if current_machine:
                            result["machines"].append(current_machine)
                        
                        machine_type = first_cell.replace('type:', '').replace('tipo:', '').replace('TYPE:', '').strip()
                        if machine_type:
                            current_machine = {
                                "type": machine_type,
                                "description": "",
                                "impostos": {
                                    "PIS_COFINS": "INCL",
                                    "IPI": "ISENTO",
                                    "ICMS": "12%",
                                    "PRAZO": "45 a 60 dias",
                                    "FRETE": "FOB/Cabreúva/SP"
                                },
                                "configuracoes_instalacao": [],
                                "baseValues": {},
                                "options": [],
                                "voltages": [],
                                "dimensions": {},
                                "peso": 0,
                                "categoria": ""
                            }
                            print(f"  🆕 Nova máquina: {machine_type}")
                    
                    # Capacidade e valor
                    elif current_machine is not None and len(row) >= 2:
                        capacity = str(row.iloc[0]).strip()
                        if capacity and capacity.lower() not in ['', 'capacity', 'capacidade', 'value', 'valor']:
                            try:
                                value = float(row.iloc[1]) if len(row) > 1 and pd.notna(row.iloc[1]) else 0
                                current_machine["baseValues"][capacity] = value
                                print(f"    📏 {capacity}: {value}")
                            except (ValueError, TypeError):
                                pass
                
                # Adicionar última máquina
                if current_machine:
                    result["machines"].append(current_machine)
            
            print(f"✅ Máquinas processadas: {len(result['machines'])}")
            
        except Exception as e:
            print(f"⚠️  Erro ao processar machines: {e}")
            traceback.print_exc()
    
    def _process_materials_sheet(self, df: pd.DataFrame, result: Dict[str, Any]) -> None:
        """Processa sheet de materiais"""
        try:
            print(f"📦 Processando materials ({len(df)} linhas)")
            
            # Verificar cabeçalho
            has_header = False
            if len(df) > 0:
                first_vals = [str(val).lower() for val in df.iloc[0].values[:3] if pd.notna(val)]
                header_keywords = ['material', 'item', 'key', 'código', 'code', 'nome']
                if any(any(kw in val for kw in header_keywords) for val in first_vals):
                    has_header = True
            
            start_row = 1 if has_header else 0
            
            for idx in range(start_row, len(df)):
                row = df.iloc[idx]
                
                if len(row) < 1 or pd.isna(row.iloc[0]):
                    continue
                
                key = str(row.iloc[0]).strip()
                if not key:
                    continue
                
                # Encontrar valor
                value = None
                for col_idx in range(1, min(5, len(row))):
                    if pd.notna(row.iloc[col_idx]):
                        cell_val = row.iloc[col_idx]
                        try:
                            if isinstance(cell_val, (int, float)):
                                value = float(cell_val)
                            elif isinstance(cell_val, str):
                                value = float(cell_val.replace(',', '.'))
                            else:
                                value = cell_val
                            break
                        except (ValueError, TypeError):
                            value = cell_val
                            break
                
                if value is None:
                    continue
                
                # Unidade e descrição
                unit = "un"
                description = ""
                
                for col_idx in range(2, min(6, len(row))):
                    if pd.notna(row.iloc[col_idx]):
                        cell_str = str(row.iloc[col_idx]).strip()
                        if len(cell_str) <= 10 and cell_str.isalpha():
                            unit = cell_str
                        elif description == "":
                            description = cell_str
                
                result["materials"][key] = {
                    "value": value,
                    "unit": unit,
                    "description": description,
                    "category": ""
                }
                
                print(f"  ✅ {key}: {value} {unit} ({description[:20]})")
            
            print(f"✅ Materiais processados: {len(result['materials'])}")
            
        except Exception as e:
            print(f"⚠️  Erro ao processar materials: {e}")
            traceback.print_exc()
    
    def _process_empresas_sheet(self, df: pd.DataFrame, result: Dict[str, Any]) -> None:
        """Processa sheet de empresas"""
        try:
            print(f"🏢 Processando empresas ({len(df)} linhas)")
            
            # Verificar formato
            if len(df.columns) < 2:
                print("⚠️  Sheet de empresas precisa de pelo menos 2 colunas")
                return
            
            # Detectar se tem cabeçalho
            has_header = False
            if len(df) > 0:
                first_row = [str(val).lower() for val in df.iloc[0].values[:2] if pd.notna(val)]
                header_keywords = ['sigla', 'code', 'código', 'empresa', 'company', 'nome', 'name']
                if any(any(kw in val for kw in header_keywords) for val in first_row):
                    has_header = True
            
            start_row = 1 if has_header else 0
            
            empresas_count = 0
            for idx in range(start_row, len(df)):
                row = df.iloc[idx]
                
                if len(row) < 2 or pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == '':
                    continue
                
                sigla = str(row.iloc[0]).strip()
                
                # Encontrar nome da empresa
                empresa_nome = ""
                for col_idx in range(1, min(5, len(row))):
                    if pd.notna(row.iloc[col_idx]):
                        empresa_nome = str(row.iloc[col_idx]).strip()
                        break
                
                if empresa_nome:
                    # Formato simplificado: {sigla: empresa_nome}
                    result["empresas"].append({sigla: empresa_nome})
                    empresas_count += 1
                    print(f"  ✅ {sigla}: {empresa_nome[:30]}")
            
            print(f"✅ Empresas processadas: {empresas_count}")
            
        except Exception as e:
            print(f"⚠️  Erro ao processar empresas: {e}")
            traceback.print_exc()
    
    def _process_equipamentos_sheet(self, df: pd.DataFrame, result: Dict[str, Any]) -> None:
        """Processa sheet de equipamentos"""
        try:
            print(f"🔧 Processando equipamentos ({len(df)} linhas)")
            
            current_tipo = None
            current_equipamento = {}
            
            for idx, row in df.iterrows():
                if len(row) < 1:
                    continue
                
                first_cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
                
                # Nova linha de tipo
                if any(keyword in first_cell.lower() for keyword in ['tipo', 'type', 'equipamento:', 'equipment:']):
                    if current_tipo and current_equipamento:
                        result["banco_equipamentos"][current_tipo] = current_equipamento
                    
                    # Extrair tipo
                    tipo = first_cell
                    for prefix in ['tipo:', 'type:', 'equipamento:', 'equipment:']:
                        if prefix in tipo.lower():
                            tipo = tipo.lower().split(prefix)[-1].strip()
                            break
                    
                    tipo = tipo.strip().title()
                    if tipo:
                        current_tipo = tipo
                        current_equipamento = {
                            "descricao": tipo,
                            "valores_padrao": {},
                            "unidade_valor": "un",
                            "dimensoes": [],
                            "categoria": "",
                            "observacoes": ""
                        }
                        print(f"  🆕 Tipo equipamento: {tipo}")
                
                # Linha de valores
                elif current_tipo and len(row) >= 2:
                    dimensao = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                    valor = row.iloc[1] if len(row) > 1 and pd.notna(row.iloc[1]) else None
                    
                    if dimensao and dimensao.lower() not in ['dimensão', 'dimension', 'size', 'tamanho'] and valor is not None:
                        try:
                            if isinstance(valor, (int, float)):
                                num_valor = float(valor)
                            elif isinstance(valor, str):
                                num_valor = float(valor.replace(',', '.'))
                            else:
                                num_valor = valor
                            
                            current_equipamento["valores_padrao"][dimensao] = num_valor
                            print(f"    📏 {dimensao}: {num_valor}")
                        except (ValueError, TypeError):
                            print(f"    ⚠️  Valor inválido para {dimensao}: {valor}")
            
            # Adicionar último equipamento
            if current_tipo and current_equipamento:
                result["banco_equipamentos"][current_tipo] = current_equipamento
            
            print(f"✅ Equipamentos processados: {len(result['banco_equipamentos'])} tipos")
            
        except Exception as e:
            print(f"⚠️  Erro ao processar equipamentos: {e}")
            traceback.print_exc()
    
    def _process_dutos_sheet(self, df: pd.DataFrame, result: Dict[str, Any]) -> None:
        """Processa sheet de dutos"""
        try:
            print(f"🟦 Processando dutos ({len(df)} linhas)")
            
            dutos = []
            current_duto = None
            
            for idx, row in df.iterrows():
                if len(row) < 1:
                    continue
                
                # Pular linhas totalmente vazias
                if all(pd.isna(cell) for cell in row):
                    continue
                
                first_cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
                
                # Identificar novo duto
                is_novo_duto = False
                if 'tipo:' in first_cell.lower() or 'type:' in first_cell.lower() or 'duto:' in first_cell.lower():
                    is_novo_duto = True
                elif idx == 0 and any(word in first_cell.lower() for word in ['tipo', 'type', 'duto']):
                    is_novo_duto = False  # Provavelmente cabeçalho
                elif current_duto is None:
                    is_novo_duto = True
                
                if is_novo_duto:
                    # Salvar duto anterior
                    if current_duto:
                        dutos.append(current_duto)
                    
                    # Extrair tipo
                    tipo = first_cell
                    for prefix in ['tipo:', 'type:', 'duto:']:
                        if prefix in tipo.lower():
                            tipo = tipo.lower().split(prefix)[-1].strip().title()
                            break
                    
                    if tipo:
                        current_duto = {
                            "type": tipo,
                            "valor": 0,
                            "descricao": "",
                            "categoria": "",
                            "unidade": "m²",
                            "opcionais": []
                        }
                        print(f"  🆕 Duto: {tipo}")
                
                # Processar linha de dados
                elif current_duto is not None:
                    desc_cell = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                    
                    # Valor base
                    if 'valor' in desc_cell.lower() and len(row) >= 2 and pd.notna(row.iloc[1]):
                        try:
                            valor = row.iloc[1]
                            if isinstance(valor, (int, float)):
                                current_duto["valor"] = float(valor)
                            elif isinstance(valor, str):
                                current_duto["valor"] = float(valor.replace(',', '.'))
                            print(f"    💰 Valor base: {current_duto['valor']}")
                        except (ValueError, TypeError):
                            print(f"    ⚠️  Valor inválido: {valor}")
                    
                    # Opcional
                    elif ('opcional' in desc_cell.lower() or 'adicional' in desc_cell.lower()) and len(row) >= 2:
                        valor = row.iloc[1] if len(row) > 1 and pd.notna(row.iloc[1]) else None
                        
                        if valor is not None:
                            try:
                                if isinstance(valor, (int, float)):
                                    num_valor = float(valor)
                                elif isinstance(valor, str):
                                    num_valor = float(valor.replace(',', '.'))
                                else:
                                    num_valor = valor
                                
                                opcional_id = len(current_duto["opcionais"]) + 1
                                opcional_nome = desc_cell.replace('opcional:', '').replace('adicional:', '').strip()
                                
                                current_duto["opcionais"].append({
                                    "id": opcional_id,
                                    "nome": opcional_nome,
                                    "value": num_valor,
                                    "descricao": ""
                                })
                                
                                print(f"    🔘 Opcional: {opcional_nome} = {num_valor}")
                            except (ValueError, TypeError):
                                print(f"    ⚠️  Valor inválido para opcional: {valor}")
                    
                    # Descrição
                    elif desc_cell and not any(word in desc_cell.lower() for word in ['valor', 'opcional', 'adicional']):
                        if not current_duto["descricao"]:
                            current_duto["descricao"] = desc_cell
                            print(f"    📝 Descrição: {desc_cell}")
            
            # Adicionar último duto
            if current_duto:
                dutos.append(current_duto)
            
            result["dutos"] = dutos
            print(f"✅ Dutos processados: {len(dutos)} tipos")
            
        except Exception as e:
            print(f"⚠️  Erro ao processar dutos: {e}")
            traceback.print_exc()
            result["dutos"] = []
    
    def _process_tubos_sheet(self, df: pd.DataFrame, result: Dict[str, Any]) -> None:
        """Processa sheet de tubos"""
        try:
            print(f"🔵 Processando tubos ({len(df)} linhas, {len(df.columns)} colunas)")
            
            tubos = []
            
            # Detectar cabeçalho
            has_header = False
            if len(df) > 0:
                first_row_vals = []
                for i in range(min(3, len(df.columns))):
                    if i < len(df.columns):
                        val = df.iloc[0, i]
                        if pd.notna(val):
                            first_row_vals.append(str(val).lower())
                
                header_keywords = ['polegadas', 'mm', 'milímetros', 'valor', 'price', 'preço', 'descrição', 'description']
                if any(any(kw in val for kw in header_keywords) for val in first_row_vals):
                    has_header = True
                    print("📋 Cabeçalho detectado")
            
            start_row = 1 if has_header else 0
            
            for idx in range(start_row, len(df)):
                row = df.iloc[idx]
                
                # Pular linhas vazias
                if all(pd.isna(cell) for cell in row):
                    continue
                
                polegadas = None
                mm = None
                valor = None
                descricao = ""
                
                # Procurar por padrões em todas as colunas
                for col_idx in range(min(5, len(row))):
                    if pd.isna(row.iloc[col_idx]):
                        continue
                    
                    cell = row.iloc[col_idx]
                    cell_str = str(cell).strip()
                    
                    # Identificar polegadas (fracionárias ou com aspas)
                    if polegadas is None:
                        if '/' in cell_str or '"' in cell_str or "''" in cell_str:
                            polegadas = cell_str.replace('"', '').replace("''", '').strip()
                            print(f"    📍 Col {col_idx}: Polegadas = {polegadas}")
                        elif any(word in cell_str.lower() for word in ['polegada', 'inch', 'in']):
                            # Pode ser rótulo, ver próxima coluna
                            if col_idx + 1 < len(row) and pd.notna(row.iloc[col_idx + 1]):
                                polegadas = str(row.iloc[col_idx + 1]).strip()
                                print(f"    📍 Col {col_idx}+1: Polegadas = {polegadas}")
                    
                    # Identificar mm
                    if mm is None:
                        if 'mm' in cell_str.lower():
                            mm_str = cell_str.lower().replace('mm', '').strip()
                            try:
                                mm = float(mm_str.replace(',', '.'))
                                print(f"    📏 Col {col_idx}: mm = {mm}")
                            except:
                                pass
                        elif col_idx == 1 and polegadas is not None:  # Segunda coluna pode ser mm
                            try:
                                mm = float(cell_str.replace(',', '.'))
                                print(f"    📏 Col {col_idx}: mm = {mm} (assumido)")
                            except:
                                pass
                    
                    # Identificar valor
                    if valor is None:
                        if any(char in cell_str for char in ['R$', '$', '€', '£']):
                            valor_str = cell_str.replace('R$', '').replace('$', '').replace('€', '').replace('£', '').strip()
                            try:
                                valor = float(valor_str.replace(',', '.'))
                                print(f"    💰 Col {col_idx}: Valor = {valor}")
                            except:
                                pass
                        elif col_idx == 2 and polegadas is not None and mm is not None:  # Terceira coluna pode ser valor
                            try:
                                valor = float(cell_str.replace(',', '.'))
                                print(f"    💰 Col {col_idx}: Valor = {valor} (assumido)")
                            except:
                                pass
                    
                    # Descrição (qualquer texto mais longo)
                    if not descricao and len(cell_str) > 10 and cell_str != cell_str.upper():
                        descricao = cell_str
                
                # Se não identificou por padrão, usar posições padrão
                if polegadas is None and len(row) >= 1 and pd.notna(row.iloc[0]):
                    polegadas = str(row.iloc[0]).strip()
                    print(f"    📍 Polegadas (padrão): {polegadas}")
                
                if mm is None and len(row) >= 2 and pd.notna(row.iloc[1]):
                    try:
                        mm = float(str(row.iloc[1]).replace(',', '.'))
                        print(f"    📏 mm (padrão): {mm}")
                    except:
                        mm = 0
                
                if valor is None and len(row) >= 3 and pd.notna(row.iloc[2]):
                    try:
                        valor_str = str(row.iloc[2]).replace('R$', '').replace('$', '').strip()
                        valor = float(valor_str.replace(',', '.'))
                        print(f"    💰 Valor (padrão): {valor}")
                    except:
                        valor = 0
                
                # Validar e adicionar tubo
                if polegadas and polegadas != "":
                    # Converter mm para número se necessário
                    if isinstance(mm, str):
                        try:
                            mm = float(mm.replace(',', '.'))
                        except:
                            mm = 0
                    
                    if mm is None:
                        mm = 0
                    
                    if valor is None:
                        valor = 0
                    
                    tubo = {
                        "polegadas": polegadas,
                        "mm": mm,
                        "valor": valor,
                        "descricao": descricao
                    }
                    
                    tubos.append(tubo)
                    print(f"  ✅ Tubo {idx}: {polegadas}'' → {mm}mm, R${valor}")
                else:
                    print(f"  ⚠️  Linha {idx} ignorada (sem polegadas)")
            
            result["tubos"] = tubos
            print(f"✅ Tubos processados: {len(tubos)}")
            
        except Exception as e:
            print(f"⚠️  Erro ao processar tubos: {e}")
            traceback.print_exc()
            result["tubos"] = []
            
    def _validate_structure(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Valida estrutura básica dos dados"""
        errors = []
        
        required_sections = [
            'constants', 
            'machines', 
            'materials', 
            'empresas', 
            'banco_equipamentos',
            'dutos',
            'tubos'
        ]
        
        for section in required_sections:
            if section not in data:
                errors.append(f"Seção '{section}' não encontrada")
        
        # Verificar tipos
        if 'dutos' in data and not isinstance(data['dutos'], list):
            errors.append("'dutos' deve ser um array")
        
        if 'tubos' in data and not isinstance(data['tubos'], list):
            errors.append("'tubos' deve ser um array")
        
        return {
            "valid": len(errors) == 0,
            "errors": errors
        }
    
    def json_to_excel(self, system_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Converte JSON do sistema para arquivo Excel
        """
        try:
            print(f"📤 Convertendo JSON para Excel...")
            print(f"📊 Estrutura recebida: {list(system_data.keys())}")
            
            # Validar dados
            validation = self._validate_structure(system_data)
            if not validation["valid"]:
                error_msg = f"Dados inválidos: {', '.join(validation['errors'])}"
                print(f"❌ {error_msg}")
                return {
                    "success": False,
                    "error": error_msg
                }
            
            # Criar workbook
            wb = Workbook()
            
            # ========== CONSTANTS ==========
            if system_data.get("constants"):
                ws_constants = wb.create_sheet("Constants")
                ws_constants.append(["Key", "Value", "Description", "Unit", "Category"])
                
                for key, const in system_data["constants"].items():
                    if isinstance(const, dict):
                        ws_constants.append([
                            key,
                            const.get("value", ""),
                            const.get("description", ""),
                            const.get("unit", ""),
                            const.get("category", "")
                        ])
                
                print(f"✅ Sheet Constants: {len(system_data['constants'])} itens")
            
            # ========== MACHINES ==========
            if system_data.get("machines"):
                ws_machines = wb.create_sheet("Machines")
                ws_machines.append([
                    "Type", "Description", "Category", "Weight",
                    "PIS_COFINS", "IPI", "ICMS", "PRAZO", "FRETE"
                ])
                
                for machine in system_data["machines"]:
                    impostos = machine.get("impostos", {})
                    ws_machines.append([
                        machine.get("type", ""),
                        machine.get("description", ""),
                        machine.get("categoria", ""),
                        machine.get("peso", 0),
                        impostos.get("PIS_COFINS", "INCL"),
                        impostos.get("IPI", "ISENTO"),
                        impostos.get("ICMS", "12%"),
                        impostos.get("PRAZO", "45 a 60 dias"),
                        impostos.get("FRETE", "FOB/Cabreúva/SP")
                    ])
                
                print(f"✅ Sheet Machines: {len(system_data['machines'])} máquinas")
            
            # ========== MATERIALS ==========
            if system_data.get("materials"):
                ws_materials = wb.create_sheet("Materials")
                ws_materials.append(["Key", "Value", "Unit", "Description", "Category"])
                
                for key, material in system_data["materials"].items():
                    if isinstance(material, dict):
                        ws_materials.append([
                            key,
                            material.get("value", ""),
                            material.get("unit", ""),
                            material.get("description", ""),
                            material.get("category", "")
                        ])
                
                print(f"✅ Sheet Materials: {len(system_data['materials'])} itens")
            
            # ========== EMPRESAS ==========
            if system_data.get("empresas"):
                ws_empresas = wb.create_sheet("Empresas")
                ws_empresas.append(["Sigla", "Nome", "CNPJ", "Endereço", "Contato", "Email", "Telefone", "Responsável"])
                
                for empresa in system_data["empresas"]:
                    if isinstance(empresa, dict):
                        for sigla, dados in empresa.items():
                            if isinstance(dados, dict):
                                ws_empresas.append([
                                    sigla,
                                    dados.get("nome", ""),
                                    dados.get("cnpj", ""),
                                    dados.get("endereco", ""),
                                    dados.get("contato", ""),
                                    dados.get("email", ""),
                                    dados.get("telefone", ""),
                                    dados.get("responsavel", "")
                                ])
                            else:
                                # Formato simplificado {sigla: nome}
                                ws_empresas.append([sigla, dados, "", "", "", "", "", ""])
                
                print(f"✅ Sheet Empresas: {len(system_data['empresas'])} empresas")
            
            # ========== EQUIPAMENTOS ==========
            if system_data.get("banco_equipamentos"):
                ws_equipamentos = wb.create_sheet("Equipamentos")
                ws_equipamentos.append(["Tipo", "Descrição", "Dimensão", "Valor", "Unidade"])
                
                for tipo, dados in system_data["banco_equipamentos"].items():
                    descricao = dados.get("descricao", tipo)
                    unidade = dados.get("unidade_valor", "un")
                    valores = dados.get("valores_padrao", {})
                    
                    if valores:
                        for dimensao, valor in valores.items():
                            ws_equipamentos.append([
                                tipo,
                                descricao,
                                dimensao,
                                valor,
                                unidade
                            ])
                    else:
                        ws_equipamentos.append([tipo, descricao, "", "", unidade])
                
                print(f"✅ Sheet Equipamentos: {len(system_data['banco_equipamentos'])} tipos")
            
            # ========== DUTOS ==========
            if system_data.get("dutos"):
                ws_dutos = wb.create_sheet("Dutos")
                ws_dutos.append(["Tipo", "Valor Base", "Descrição", "Categoria", "Unidade"])
                
                for duto in system_data["dutos"]:
                    ws_dutos.append([
                        duto.get("type", ""),
                        duto.get("valor", 0),
                        duto.get("descricao", ""),
                        duto.get("categoria", ""),
                        duto.get("unidade", "m²")
                    ])
                    
                    # Opcionais em sub-sheet
                    if duto.get("opcionais"):
                        tipo_limpo = duto.get("type", "Desconhecido").replace("/", "_").replace("\\", "_")
                        sheet_name = f"Opc_{tipo_limpo[:25]}"
                        if sheet_name not in wb.sheetnames:
                            ws_opcionais = wb.create_sheet(sheet_name)
                            ws_opcionais.append(["Tipo Duto", "ID", "Nome Opcional", "Valor", "Descrição"])
                        else:
                            ws_opcionais = wb[sheet_name]
                        
                        for opcional in duto["opcionais"]:
                            ws_opcionais.append([
                                duto.get("type", ""),
                                opcional.get("id", ""),
                                opcional.get("nome", ""),
                                opcional.get("value", 0),
                                opcional.get("descricao", "")
                            ])
                
                print(f"✅ Sheet Dutos: {len(system_data['dutos'])} tipos")
            
            # ========== TUBOS ==========
            if system_data.get("tubos"):
                ws_tubos = wb.create_sheet("Tubos")
                ws_tubos.append(["Polegadas", "Milímetros (mm)", "Valor (R$)", "Descrição"])
                
                for tubo in system_data["tubos"]:
                    ws_tubos.append([
                        tubo.get("polegadas", ""),
                        tubo.get("mm", 0),
                        tubo.get("valor", 0),
                        tubo.get("descricao", "")
                    ])
                
                print(f"✅ Sheet Tubos: {len(system_data['tubos'])} tipos")
            
            # Remover sheet padrão vazio
            if 'Sheet' in wb.sheetnames:
                std = wb['Sheet']
                wb.remove(std)
            
            # Salvar em buffer
            from io import BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Converter para base64
            excel_bytes = buffer.getvalue()
            excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
            
            # Nome do arquivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"sistema_export_{timestamp}.xlsx"
            
            print("\n" + "="*50)
            print("✅ EXCEL GERADO COM SUCESSO")
            print("="*50)
            print(f"📄 Arquivo: {filename}")
            print(f"📊 Sheets: {wb.sheetnames}")
            print(f"📈 Tamanho: {len(excel_bytes)} bytes")
            print("="*50)
            
            return {
                "success": True,
                "filename": filename,
                "data": excel_base64,
                "sheets": wb.sheetnames,
                "message": "Excel exportado com sucesso"
            }
            
        except Exception as e:
            print(f"❌ Erro crítico ao gerar Excel: {e}")
            traceback.print_exc()
            return {
                "success": False,
                "error": f"Erro na geração do Excel: {str(e)}"
            }

# Instância global do conversor
converter = ExcelConverter()